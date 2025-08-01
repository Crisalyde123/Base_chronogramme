import openpyxl
import pandas as pd
import unicodedata
import logging
import warnings
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

logger = logging.getLogger(__name__)


def normalize_text(text: str) -> str:
    """Return a lowercase, accent-free version of `text`."""
    if text is None:
        return ""
    text = str(text).strip()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    return text.lower()


def load_header_mapping(path: Path) -> Dict[str, str]:
    """Load header mapping CSV into a dictionary with normalized keys/values."""
    mapping: Dict[str, str] = {}
    if not path.exists():
        logger.warning("Header mapping file %s not found", path)
        return mapping
    df = pd.read_csv(path)
    for _, row in df.iterrows():
        orig = normalize_text(row[0])
        std = normalize_text(row[1])
        mapping[orig] = std
    return mapping


def detect_main_sheet(file: Path) -> str:
    """Return the sheet name with the most non-empty cells."""
    xls = pd.ExcelFile(file)
    best_sheet = xls.sheet_names[0]
    max_count = -1
    for sheet in xls.sheet_names:
        df = xls.parse(sheet, header=None)
        count = int(df.count().sum())
        if count > max_count:
            max_count = count
            best_sheet = sheet
    logger.debug("%s -> main sheet '%s'", file.name, best_sheet)
    return best_sheet


def find_header_row(df: pd.DataFrame, keywords: Iterable[str]) -> int:
    """Return index of row likely containing column headers."""
    for idx, row in df.iterrows():
        values = [normalize_text(v) for v in row]
        matches = sum(any(k in val for val in values) for k in keywords)
        if matches >= 2:
            return idx
    return 0


def extract_table(file: Path, sheet: str) -> pd.DataFrame:
    """Extract main data table from *sheet* of *file*."""
    df_raw = pd.read_excel(file, sheet_name=sheet, header=None)
    header_row = find_header_row(
        df_raw, ["emetteur", "destinataire", "recepteur", "modalite", "type", "nature"]
    )
    df = pd.read_excel(file, sheet_name=sheet, header=header_row)
    df.dropna(axis=1, how="all", inplace=True)
    return df


def standardize_headers(df: pd.DataFrame, mapping: Dict[str, str]) -> None:
    """Rename DataFrame columns using mapping after normalization."""
    new_cols: List[str] = []
    for col in df.columns:
        norm = normalize_text(col)
        new_cols.append(mapping.get(norm, norm))
    df.columns = new_cols


def canonical_column(name: str) -> str | None:
    """Map normalized column name to canonical target column."""
    name = normalize_text(name)
    if "emetteur" in name:
        return "emetteur"
    if "destinataire" in name or "recepteur" in name:
        return "recepteur"
    if "modalite" in name:
        return "modalite"
    if "type" in name or "nature" in name:
        return "type_inject"
    return None


def enrich_mapping_values(input_dir: Path, header_map_path: Path, values_path: Path, limit: int = 3) -> None:
    """Process Excel files in *input_dir* and update mapping_values.csv."""
    mapping = load_header_mapping(header_map_path)
    values_df = pd.read_csv(values_path)
    existing = set(zip(values_df["Colonne"], values_df["Valeur brute"]))
    added_lines = 0
    cols_inspected = 0

    files = sorted(input_dir.glob("*.xlsx"))[:limit]
    for file in files:
        sheet = detect_main_sheet(file)
        df = extract_table(file, sheet)
        standardize_headers(df, mapping)
        for col in list(df.columns):
            canon = canonical_column(col)
            if canon is None:
                continue
            cols_inspected += 1
            uniques = pd.Series(df[col].dropna().unique()).astype(str).str.strip()
            for val in uniques:
                key = (canon, val)
                if key not in existing:
                    values_df.loc[len(values_df)] = [canon, val, ""]
                    existing.add(key)
                    added_lines += 1
    values_df.drop_duplicates(inplace=True)
    values_df.to_csv(values_path, index=False)
    logger.info("%d columns inspected", cols_inspected)
    logger.info("%d new raw values added", added_lines)
    logger.info("%d total lines now in %s", len(values_df), values_path)


# --- Functions specific to openpyxl-based processing ---

def detect_main_sheet_openpyxl(workbook_path: Path) -> openpyxl.worksheet.worksheet.Worksheet:
    """Return the worksheet with the highest number of non-empty cells."""
    with warnings.catch_warnings():
        # ignore unsupported Data Validation extension warnings from openpyxl
        warnings.simplefilter("ignore", UserWarning)
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
    best_sheet = wb.worksheets[0]
    max_count = -1
    for ws in wb.worksheets:
        count = sum(
            1
            for row in ws.iter_rows(values_only=True)
            for cell in row
            if cell not in (None, "")
        )
        if count > max_count:
            best_sheet = ws
            max_count = count
    logger.info("Selected sheet '%s' with %d cells from %s", best_sheet.title, max_count, workbook_path)
    return best_sheet


def find_data_table(sheet: openpyxl.worksheet.worksheet.Worksheet) -> Tuple[int, int, int]:
    """Return header row index and start/end columns of the data table."""
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [cell for cell in row if cell not in (None, "")]
        if len(values) >= 3:
            first_col = next(i for i, c in enumerate(row, start=1) if c not in (None, ""))
            last_col = len(row) - next(i for i, c in enumerate(reversed(row), start=1) if c not in (None, "")) + 1
            logger.info("Header row detected at line %d", idx)
            return idx, first_col, last_col
    raise ValueError("No header row found")


def extract_headers(
    sheet: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    first_col: int,
    last_col: int,
) -> List[str]:
    """Return header values from ``sheet`` between ``first_col`` and ``last_col``."""
    cells = sheet.iter_rows(
        min_row=header_row,
        max_row=header_row,
        min_col=first_col,
        max_col=last_col,
        values_only=True,
    )
    headers = []
    for row in cells:
        for value in row:
            if value not in (None, ""):
                headers.append(str(value).strip())
    return headers


def update_mapping_headers(input_dir: Path, mapping_csv: Path, max_files: int = 3) -> Tuple[int, int]:
    """Update mapping CSV with headers extracted from Excel files.

    Returns (total_headers_found, new_headers_added).
    """
    excel_files = [p for p in input_dir.iterdir() if p.suffix.lower() in {".xlsx", ".xls"}]
    excel_files.sort()
    if not excel_files:
        logger.warning("No Excel files found in %s", input_dir)
        return 0, 0
    excel_files = excel_files[:max_files]

    all_headers: List[str] = []
    for file in excel_files:
        sheet = detect_main_sheet_openpyxl(file)
        header_row, first_col, last_col = find_data_table(sheet)
        headers = extract_headers(sheet, header_row, first_col, last_col)
        logger.info("%s -> %d headers", file, len(headers))
        all_headers.extend(headers)

    total_headers = len(all_headers)
    headers_set = {h for h in all_headers if h}

    if mapping_csv.exists() and mapping_csv.stat().st_size > 0:
        df = pd.read_csv(mapping_csv)
    else:
        df = pd.DataFrame(columns=["En-tête original", "En-tête standard"])

    existing = set(df["En-tête original"].astype(str))
    new_entries = [h for h in headers_set if h not in existing]
    for header in new_entries:
        df.loc[len(df)] = [header, ""]

    if new_entries:
        df.to_csv(mapping_csv, index=False)
    logger.info("%d headers extracted, %d new entries added", total_headers, len(new_entries))
    return total_headers, len(new_entries)
