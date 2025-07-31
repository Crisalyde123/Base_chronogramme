import logging
from pathlib import Path
from typing import Union

from openpyxl import load_workbook, Workbook
import warnings

logger = logging.getLogger(__name__)

KEYWORDS = {"chronogramme", "exercice"}


def _count_non_empty(ws) -> int:
    """Return number of non-empty cells in worksheet."""
    count = 0
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell not in (None, ""):
                count += 1
    return count


def _contains_keyword(ws) -> bool:
    title = ws.title.lower()
    if any(k in title for k in KEYWORDS):
        return True
    # check first few cells
    for row in ws.iter_rows(min_row=1, max_row=5, max_col=5, values_only=True):
        for cell in row:
            if isinstance(cell, str) and any(k in cell.lower() for k in KEYWORDS):
                return True
    return False


def _ai_select_sheet(candidates):
    """Use OpenAI to select sheet from candidates. Return worksheet."""
    try:
        import openai
    except Exception as exc:  # pragma: no cover - optional dependency
        raise RuntimeError("openai not available") from exc

    if not openai.api_key:
        raise RuntimeError("OPENAI_API_KEY not configured")

    prompt = [
        {
            "role": "system",
            "content": (
                "Parmi les feuilles suivantes d'un classeur Excel, laquelle contient"
                " probablement un chronogramme d'exercice ? Réponds uniquement par le"
                " nom de la feuille."
            ),
        }
    ]

    for ws, _kw, _count in candidates:
        sample_lines = []
        for row in ws.iter_rows(min_row=1, max_row=5, max_col=5, values_only=True):
            sample_lines.append(
                "\t".join("" if c is None else str(c) for c in row)
            )
        sample = "\n".join(sample_lines)
        prompt.append({"role": "user", "content": f"Feuille: {ws.title}\n{sample}"})

    response = openai.ChatCompletion.create(model="gpt-4-turbo", messages=prompt)
    name = response.choices[0].message.content.strip()
    for ws, _kw, _count in candidates:
        if ws.title.lower() == name.lower():
            return ws
    # if not matched, return first
    return candidates[0][0]


def detect_main_sheet(workbook: Union[str, Path, Workbook]):
    """Detect the worksheet containing the main chronogram table."""
    if isinstance(workbook, (str, Path)):
        with warnings.catch_warnings():
            # ignore unsupported Data Validation extension warnings from openpyxl
            warnings.simplefilter("ignore", UserWarning)
            wb = load_workbook(workbook, data_only=True)
    elif isinstance(workbook, Workbook):
        wb = workbook
    else:
        raise TypeError("workbook must be path or Workbook")

    sheet_info = []
    for ws in wb.worksheets:
        non_empty = _count_non_empty(ws)
        total = ws.max_row * ws.max_column if ws.max_row and ws.max_column else 0
        density = non_empty / total if total else 0
        has_kw = _contains_keyword(ws)
        logger.debug(
            "Sheet %s: %d non-empty cells, density %.3f, keyword %s",
            ws.title,
            non_empty,
            density,
            has_kw,
        )
        sheet_info.append((ws, has_kw, non_empty, density))

    # filter by keyword if any
    sheets_with_kw = [info for info in sheet_info if info[1]]
    candidates = sheets_with_kw if sheets_with_kw else sheet_info

    # pick with highest non-empty cell count
    max_count = max(info[2] for info in candidates) if candidates else 0
    top_candidates = [info for info in candidates if info[2] == max_count]

    if len(top_candidates) == 1:
        ws, has_kw, count, _ = top_candidates[0]
        logger.info("Selected sheet '%s' (%d non-empty cells)", ws.title, count)
        return ws.title

    # multiple candidates -> try AI
    try:
        ws = _ai_select_sheet(top_candidates)
        logger.info("Selected sheet '%s' via AI", ws.title)
        return ws.title
    except Exception as exc:  # pragma: no cover - fallback
        ws, has_kw, count, _ = top_candidates[0]
        logger.warning(
            "AI selection failed (%s). Defaulting to sheet '%s'", exc, ws.title
        )
        return ws.title
