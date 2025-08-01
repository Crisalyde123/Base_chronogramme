import sys
from pathlib import Path

import openpyxl

# allow imports from project root
sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from src.excel_parser import detect_main_sheet, _count_non_empty, _contains_keyword

INPUTS_DIR = Path(__file__).resolve().parents[1] / "data" / "inputs"


def test_detect_main_sheet_prefers_keyword():
    path = INPUTS_DIR / "test bonne feuille et trcabilité feuille.xlsx"
    assert detect_main_sheet(path) == "Chronogramme"


def test_detect_main_sheet_highest_density():
    path = INPUTS_DIR / "Kit intermédiaire.xlsx"
    assert detect_main_sheet(path) == "Chrono"


def test_detect_main_sheet_tie_returns_first():
    path = INPUTS_DIR / "Chronogramme.xlsx"
    wb = openpyxl.load_workbook(path)
    wb.copy_worksheet(wb.active)
    assert detect_main_sheet(wb) == "Chronogramme "


def test_count_non_empty():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "x"
    ws["C2"] = 42
    ws["B3"] = ""
    assert _count_non_empty(ws) == 2


def test_contains_keyword_title_and_cells():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mon Chronogramme"
    assert _contains_keyword(ws)
    ws.title = "Feuille"
    ws["A1"] = "Exercice 2024"
    assert _contains_keyword(ws)
    ws.title = "Autre"
    ws["A1"] = "foo"
    assert not _contains_keyword(ws)
